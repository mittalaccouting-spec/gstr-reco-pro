import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher
from collections import defaultdict
import io

st.set_page_config(
    page_title="GSTR-2A Reco Tool · by Harsh Mevada",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="collapsed"
)

UPI_ID     = "7600480575@upi"
UPI_NAME   = "Harsh Mevada"
UPI_AMOUNT = "5"
UPI_NOTE   = "GSTR2A Reco"
UPI_LINK   = f"upi://pay?pa={UPI_ID}&pn={UPI_NAME.replace(' ','%20')}&am={UPI_AMOUNT}&cu=INR&tn={UPI_NOTE.replace(' ','%20')}"
QR_URL     = f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={UPI_LINK}&bgcolor=0d0d1a&color=ffd200&qzone=2"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@400;500;600&family=JetBrains+Mono:wght@500;700&display=swap');

*, body {{ font-family: 'DM Sans', sans-serif; }}
.stApp {{ background: #0a0a0f; min-height: 100vh; }}
.stApp::before {{
    content: ''; position: fixed; inset: 0;
    background-image: linear-gradient(rgba(255,210,0,0.025) 1px, transparent 1px),
                      linear-gradient(90deg, rgba(255,210,0,0.025) 1px, transparent 1px);
    background-size: 56px 56px; pointer-events: none; z-index: 0;
}}
.block-container {{ position: relative; z-index: 1; padding-top: 0.5rem !important; max-width: 1000px !important; }}

.hero {{ text-align: center; padding: 2rem 1rem 0.6rem 1rem; }}
.hero-badge {{
    display: inline-block; background: rgba(255,210,0,0.08);
    border: 1px solid rgba(255,210,0,0.25); color: #ffd200;
    font-size: 0.7rem; font-weight: 600; letter-spacing: 2px; text-transform: uppercase;
    padding: 0.3rem 1rem; border-radius: 100px; margin-bottom: 1rem;
}}
.hero h1 {{
    font-family: 'Syne', sans-serif; font-size: 2.8rem; font-weight: 800;
    color: #fff; letter-spacing: -1.5px; line-height: 1.1; margin: 0 0 0.4rem 0;
}}
.hero h1 span {{
    background: linear-gradient(90deg, #f7971e, #ffd200, #f7971e);
    background-size: 200%; -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    animation: shimmer 3s infinite linear;
}}
@keyframes shimmer {{ 0% {{ background-position: 0% }} 100% {{ background-position: 200% }} }}
.hero-sub  {{ color: #777; font-size: 0.95rem; margin: 0; }}
.hero-auth {{ font-size: 0.76rem; color: #444; margin-top: 0.3rem; }}
.hero-auth span {{ color: #ffd200; font-weight: 600; }}

.security-bar {{
    background: rgba(86,227,159,0.05); border: 1px solid rgba(86,227,159,0.15);
    border-radius: 10px; padding: 0.55rem 1rem; font-size: 0.79rem;
    color: #56e39f; text-align: center; margin: 0.9rem 0 1.2rem 0;
}}
.tolerance-info {{
    background: rgba(247,151,30,0.06); border: 1px solid rgba(247,151,30,0.18);
    border-radius: 10px; padding: 0.65rem 1rem; font-size: 0.81rem;
    color: #f7971e; margin: 0.8rem 0 1rem 0;
}}
.section-label {{
    font-size: 0.71rem; font-weight: 600; letter-spacing: 1.8px;
    text-transform: uppercase; color: #ffd200; margin-bottom: 0.4rem; display: block;
}}

.stat-grid {{
    display: grid; grid-template-columns: repeat(3, 1fr); gap: 0.8rem; margin: 1.2rem 0;
}}
.stat-card {{
    background: rgba(255,255,255,0.04); border-radius: 14px; padding: 1.1rem 0.8rem;
    text-align: center; border: 1px solid rgba(255,255,255,0.07);
}}
.stat-number {{ font-size: 2rem; font-weight: 700; font-family: 'JetBrains Mono', monospace; line-height: 1; }}
.stat-label  {{ font-size: 0.75rem; color: #555; margin-top: 0.25rem; }}
.c-green {{ color: #56e39f; }} .c-yellow {{ color: #ffd200; }}
.c-red   {{ color: #ff6b6b; }} .c-blue   {{ color: #74b9ff; }}

.pay-box {{
    background: linear-gradient(135deg, rgba(247,151,30,0.07), rgba(255,210,0,0.04));
    border: 1.5px solid rgba(255,210,0,0.22); border-radius: 20px;
    padding: 2rem; margin: 1rem 0 1.5rem 0;
}}
.pay-box h3 {{
    font-family: 'Syne', sans-serif; color: #fff; font-size: 1.4rem;
    font-weight: 700; margin: 0 0 0.3rem 0; text-align: center;
}}
.pay-box .sub {{ color: #666; font-size: 0.84rem; text-align: center; margin-bottom: 1.5rem; }}
.pay-inner {{
    display: flex; gap: 2.5rem; align-items: center;
    justify-content: center; flex-wrap: wrap;
}}
.pay-left {{ text-align: center; }}
.pay-amount {{ font-family: 'JetBrains Mono', monospace; font-size: 4rem; font-weight: 700; color: #ffd200; line-height: 1; }}
.pay-per {{ font-size: 0.76rem; color: #555; margin-top: 0.2rem; }}
.pay-right {{ text-align: center; }}
.qr-wrap {{
    background: #0d0d1a; border: 2px solid rgba(255,210,0,0.3);
    border-radius: 16px; padding: 12px; display: inline-block;
}}
.qr-hint {{ font-size: 0.72rem; color: #555; margin-top: 0.4rem; }}
.qr-apps {{ display: flex; justify-content: center; gap: 0.4rem; margin-top: 0.5rem; flex-wrap: wrap; }}
.qr-app {{
    background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08);
    border-radius: 6px; padding: 0.18rem 0.55rem; font-size: 0.7rem; color: #888;
}}
.refund-tag {{
    background: rgba(86,227,159,0.06); border: 1px solid rgba(86,227,159,0.15);
    border-radius: 8px; padding: 0.5rem 0.8rem; font-size: 0.75rem;
    color: #56e39f; margin-top: 1rem; text-align: center; line-height: 1.6;
}}

.stButton > button {{
    background: linear-gradient(90deg, #f7971e, #ffd200) !important;
    color: #0a0a0f !important; font-weight: 700 !important; font-size: 0.95rem !important;
    border: none !important; border-radius: 10px !important;
    padding: 0.65rem 1.5rem !important; width: 100% !important;
}}
.stDownloadButton > button {{
    background: linear-gradient(90deg, #56e39f, #11998e) !important;
    color: #fff !important; font-weight: 700 !important; font-size: 1rem !important;
    border: none !important; border-radius: 10px !important;
    padding: 0.72rem 1.5rem !important; width: 100% !important;
}}
.footer {{
    text-align: center; color: #2a2a2a; font-size: 0.72rem;
    padding: 2rem 0 1rem 0; line-height: 2;
}}
.footer span {{ color: #ffd200; }}
footer, #MainMenu, header {{ visibility: hidden; }}
div[data-testid="stFileUploader"] label {{ display: none !important; }}
</style>
""", unsafe_allow_html=True)

# ── HERO ───────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div class="hero-badge">🧾 GST Reconciliation Tool</div>
  <h1>GSTR-2A vs Books <span>Reco</span></h1>
  <p class="hero-sub">Upload your files below — get your full reconciliation summary free. Pay ₹5 only to download.</p>
  <p class="hero-auth">Created by <span>Harsh Mevada</span> · CA Tools India</p>
</div>
<div class="security-bar">
  🔒 <b>100% Private & Secure</b> — Files processed in memory only. Never saved, stored or shared with anyone.
</div>
""", unsafe_allow_html=True)

# ── UPLOAD ─────────────────────────────────────────────────────────────────────
st.markdown('<div class="tolerance-info">⚡ Smart fuzzy name matching + <b>±₹10 GST tolerance</b> per head — auto treated as matched</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    st.markdown('<span class="section-label">📥 GSTR-2A — Portal Export</span>', unsafe_allow_html=True)
    file_2a = st.file_uploader("2a", type=["xls","xlsx"], key="f2a", label_visibility="collapsed")
    if file_2a: st.success(f"✅ {file_2a.name}")

with col2:
    st.markdown('<span class="section-label">📒 Books Purchase Register — Tally Export</span>', unsafe_allow_html=True)
    file_books = st.file_uploader("books", type=["xls","xlsx"], key="fb", label_visibility="collapsed")
    if file_books: st.success(f"✅ {file_books.name}")

st.markdown("<br>", unsafe_allow_html=True)
col_run, _, _ = st.columns([1,1,1])
with col_run:
    run_btn = st.button("⚡ Run Reconciliation — Free")

# ── CORE FUNCTIONS ─────────────────────────────────────────────────────────────
def convert_xls_to_xlsx(file_bytes, filename):
    if filename.endswith('.xlsx'):
        return io.BytesIO(file_bytes)
    import xlrd
    xls_book = xlrd.open_workbook(file_contents=file_bytes)
    new_wb = openpyxl.Workbook(); new_wb.remove(new_wb.active)
    for sheet_name in xls_book.sheet_names():
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        new_ws = new_wb.create_sheet(title=sheet_name)
        for ri in range(xls_sheet.nrows):
            for ci in range(xls_sheet.ncols):
                cell = xls_sheet.cell(ri, ci)
                if cell.ctype == 3:
                    try:
                        new_ws.cell(row=ri+1, column=ci+1, value=xlrd.xldate_as_datetime(cell.value, xls_book.datemode))
                    except:
                        new_ws.cell(row=ri+1, column=ci+1, value=cell.value)
                elif cell.ctype != 0:
                    new_ws.cell(row=ri+1, column=ci+1, value=cell.value)
    buf = io.BytesIO(); new_wb.save(buf); buf.seek(0); return buf

def parse_2a(wb):
    rows = list(wb['invoice'].iter_rows(values_only=True))
    data = []
    for r in rows[3:]:
        if not r[0] or not isinstance(r[0], (int, float)): continue
        data.append({'sno':int(r[0]), 'supplier':str(r[1] or '').strip().upper(),
                     'gstin':str(r[2] or '').strip().upper(), 'period':str(r[3] or '').strip(),
                     'inv_no':str(r[5] or '').strip(), 'inv_date':str(r[8]) if r[8] else '',
                     'inv_value':float(r[9] or 0), 'taxable':float(r[10] or 0),
                     'igst':float(r[11] or 0), 'cgst':float(r[12] or 0), 'sgst':float(r[13] or 0)})
    return data

def parse_books(wb):
    rows = list(wb.active.iter_rows(values_only=True))
    CGST=[6,18,21,23]; SGST=[7,19,22,24]; IGST=[32]; data=[]
    for i, r in enumerate(rows):
        if i<9 or not r[0] or not r[1] or r[1]=='Grand Total': continue
        cgst=sum(float(r[c] or 0) for c in CGST if c<len(r))
        sgst=sum(float(r[c] or 0) for c in SGST if c<len(r))
        igst=sum(float(r[c] or 0) for c in IGST if c<len(r))
        if cgst==0 and sgst==0 and igst==0: continue
        data.append({'date':str(r[0])[:10] if r[0] else '', 'name':str(r[1] or '').strip().upper(),
                     'gross':float(r[4] or 0), 'cgst':cgst, 'sgst':sgst, 'igst':igst,
                     '_matched':False, '_id':i})
    return data

def normalize(s): return re.sub(r'[^A-Z0-9 ]','',re.sub(r'\s+',' ',s.upper().strip()))
def similarity(a,b): return SequenceMatcher(None,normalize(a),normalize(b)).ratio()

def run_matching(gst2a, books):
    by_name=defaultdict(list)
    for b in books: by_name[b['name']].append(b)
    used,exact,diff,un2a=set(),[],[],[]
    for rec in gst2a:
        best,best_score,best_d=None,0,None
        for bname,blist in by_name.items():
            sim=similarity(rec['supplier'],bname)
            if sim<0.65: continue
            for bk in blist:
                if bk['_id'] in used: continue
                cd=abs(rec['cgst']-bk['cgst']); sd=abs(rec['sgst']-bk['sgst']); gd=abs(rec['igst']-bk['igst'])
                if cd<=10 and sd<=10 and gd<=10:
                    score=sim*100+1/(cd+sd+gd+0.01)
                    if score>best_score: best_score,best,best_d=score,bk,(cd,sd,gd)
        if best:
            used.add(best['_id']); best['_matched']=True
            e={**rec,'bk_name':best['name'],'bk_date':best['date'],'bk_gross':best['gross'],
               'bk_cgst':best['cgst'],'bk_sgst':best['sgst'],'bk_igst':best['igst'],
               'cgst_diff':best_d[0],'sgst_diff':best_d[1],'igst_diff':best_d[2],'total_diff':sum(best_d)}
            (exact if sum(best_d)<0.01 else diff).append(e)
        else: un2a.append(rec)
    return exact,diff,un2a,[b for b in books if not b['_matched']]

def build_excel(res_exact,res_diff,res_un2a,res_unb,gst2a,books):
    G=PatternFill("solid",fgColor="C6EFCE"); Y=PatternFill("solid",fgColor="FFEB9C")
    RF=PatternFill("solid",fgColor="FFC7CE"); LB=PatternFill("solid",fgColor="DDEEFF")
    BH=PatternFill("solid",fgColor="1F4E79"); TH=PatternFill("solid",fgColor="375623")
    OH=PatternFill("solid",fgColor="7B6200"); RH=PatternFill("solid",fgColor="833C00")
    DH=PatternFill("solid",fgColor="7B2C2C")
    t=Side(style="thin",color="BBBBBB"); BD=Border(left=t,right=t,top=t,bottom=t)
    WH=Font(color="FFFFFF",bold=True,name="Arial",size=9); BLD=Font(bold=True,name="Arial",size=9)
    NR=Font(name="Arial",size=9); CT=Alignment(horizontal="center",vertical="center",wrap_text=True)
    LF=Alignment(horizontal="left",vertical="center")
    def sh(c,f): c.fill=f;c.font=WH;c.alignment=CT;c.border=BD
    def sc(c,f=None): c.font=NR;c.alignment=LF;c.border=BD;(setattr(c,'fill',f) if f else None)
    sf=lambda lst,k:sum(r[k] for r in lst)
    wb=openpyxl.Workbook(); wb.remove(wb.active)

    ws=wb.create_sheet("📊 Summary"); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1"); c=ws["A1"]
    c.value="GSTR-2A vs Books — Reconciliation Summary · by Harsh Mevada"
    c.font=Font(name="Arial",bold=True,size=13,color="FFFFFF"); c.fill=BH; c.alignment=CT; ws.row_dimensions[1].height=26
    for col,(label,val,fill) in enumerate([
        ("Total 2A Records",len(gst2a),BH),("Total Books Records",len(books),PatternFill("solid",fgColor="1F3864")),
        ("✅ Matched Exact",len(res_exact),TH),("✅ Matched (±₹10)",len(res_diff),OH),
        ("⚠️ Unmatched in 2A",len(res_un2a),RH),("⚠️ Unmatched in Books",len(res_unb),DH)],1):
        lc=ws.cell(row=4,column=col,value=label);lc.fill=fill;lc.font=WH;lc.alignment=CT;lc.border=BD
        vc=ws.cell(row=5,column=col,value=val)
        vc.fill=PatternFill("solid",fgColor="F2F2F2");vc.font=Font(name="Arial",bold=True,size=14);vc.alignment=CT;vc.border=BD
    ws.row_dimensions[5].height=30
    for ci,h in enumerate(["","IGST (₹)","CGST (₹)","SGST (₹)","Total GST (₹)"],1): sh(ws.cell(row=8,column=ci,value=h),BH)
    for ri,(label,igst,cgst,sgst) in enumerate([
        ("2A Total",sf(gst2a,'igst'),sf(gst2a,'cgst'),sf(gst2a,'sgst')),
        ("Books Total",sf(books,'igst'),sf(books,'cgst'),sf(books,'sgst'))],9):
        for ci,v in enumerate([label,igst,cgst,sgst,igst+cgst+sgst],1):
            cell=ws.cell(row=ri,column=ci,value=v if isinstance(v,str) else round(v,2))
            sc(cell,PatternFill("solid",fgColor="EBF5FB" if ri==9 else "FEF9E7"))
            if ci>1: cell.number_format='#,##0.00'
    dr=["Difference",round(sf(gst2a,'igst')-sf(books,'igst'),2),round(sf(gst2a,'cgst')-sf(books,'cgst'),2),round(sf(gst2a,'sgst')-sf(books,'sgst'),2)]
    dr.append(sum(dr[1:]))
    for ci,v in enumerate(dr,1):
        cell=ws.cell(row=11,column=ci,value=v if isinstance(v,str) else round(v,2))
        sc(cell,PatternFill("solid",fgColor="FCE4D6") if isinstance(v,(int,float)) and abs(v)>0.01 else G)
        if ci>1: cell.number_format='#,##0.00'
    for i in range(1,7): ws.column_dimensions[get_column_letter(i)].width=25

    def mk_matched(wb,name,data,hfill,is_diff=False):
        ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
        lc='S' if is_diff else 'P'; ws.merge_cells(f"A1:{lc}1"); c=ws["A1"]; c.value=name
        c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=hfill; c.alignment=CT
        hdrs=["Sno","2A: Supplier","2A: GSTIN","2A: Invoice No","2A: Date","2A: Inv Value","2A: Taxable",
              "2A: IGST","2A: CGST","2A: SGST","Books: Vendor","Books: Date","Books: Gross",
              "Books: IGST","Books: CGST","Books: SGST"]
        if is_diff: hdrs+=["IGST Diff","CGST Diff","SGST Diff"]
        for ci,h in enumerate(hdrs,1): sh(ws.cell(row=2,column=ci,value=h),hfill)
        rf=G if not is_diff else Y
        for ri,rec in enumerate(data,3):
            vals=[rec['sno'],rec['supplier'],rec['gstin'],rec['inv_no'],rec['inv_date'],
                  rec['inv_value'],rec['taxable'],rec['igst'],rec['cgst'],rec['sgst'],
                  rec['bk_name'],rec['bk_date'],rec['bk_gross'],rec['bk_igst'],rec['bk_cgst'],rec['bk_sgst']]
            if is_diff: vals+=[rec['igst_diff'],rec['cgst_diff'],rec['sgst_diff']]
            for ci,v in enumerate(vals,1):
                cell=ws.cell(row=ri,column=ci,value=v); sc(cell,rf)
                if ci in [6,7,8,9,10,13,14,15,16,17,18,19]: cell.number_format='#,##0.00'
        tr=len(data)+3; ws.cell(row=tr,column=1,value="TOTAL").font=BLD
        for ci,key in [(8,'igst'),(9,'cgst'),(10,'sgst'),(14,'bk_igst'),(15,'bk_cgst'),(16,'bk_sgst')]:
            c2=ws.cell(row=tr,column=ci,value=round(sum(r[key] for r in data),2))
            c2.font=BLD;c2.number_format='#,##0.00';c2.fill=LB;c2.border=BD
        ws.freeze_panes="A3"
        widths=[5,35,22,22,14,14,14,12,12,12,35,14,14,12,12,12]+([12,12,12] if is_diff else [])
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    def mk_un2a(wb,data):
        ws=wb.create_sheet("⚠️ Unmatched in 2A"); ws.sheet_view.showGridLines=False
        ws.merge_cells("A1:J1"); c=ws["A1"]; c.value="⚠️ In GSTR-2A but NOT in Books — Possible Missed ITC"
        c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=RH; c.alignment=CT
        for ci,h in enumerate(["Sno","Supplier Name","GSTIN","Period","Invoice No","Invoice Date","Invoice Value","IGST","CGST","SGST"],1):
            sh(ws.cell(row=2,column=ci,value=h),RH)
        for ri,rec in enumerate(data,3):
            for ci,v in enumerate([rec['sno'],rec['supplier'],rec['gstin'],rec['period'],rec['inv_no'],
                                    rec['inv_date'],rec['inv_value'],rec['igst'],rec['cgst'],rec['sgst']],1):
                cell=ws.cell(row=ri,column=ci,value=v); sc(cell,RF)
                if ci in [7,8,9,10]: cell.number_format='#,##0.00'
        tr=len(data)+3; ws.cell(row=tr,column=1,value="TOTAL").font=BLD
        for ci,key in [(8,'igst'),(9,'cgst'),(10,'sgst')]:
            c2=ws.cell(row=tr,column=ci,value=round(sum(r[key] for r in data),2))
            c2.font=BLD;c2.number_format='#,##0.00';c2.fill=LB;c2.border=BD
        ws.freeze_panes="A3"
        for i,w in enumerate([5,38,22,14,25,14,14,12,12,12],1): ws.column_dimensions[get_column_letter(i)].width=w

    def mk_unb(wb,data):
        ws=wb.create_sheet("⚠️ Unmatched in Books"); ws.sheet_view.showGridLines=False
        ws.merge_cells("A1:G1"); c=ws["A1"]; c.value="⚠️ In Books but NOT in 2A — Supplier May Not Have Filed GST"
        c.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); c.fill=DH; c.alignment=CT
        for ci,h in enumerate(["Vendor Name","Date","Gross Total","IGST","CGST","SGST","Total GST"],1):
            sh(ws.cell(row=2,column=ci,value=h),DH)
        for ri,rec in enumerate(data,3):
            tgst=rec['igst']+rec['cgst']+rec['sgst']
            for ci,v in enumerate([rec['name'],rec['date'],rec['gross'],rec['igst'],rec['cgst'],rec['sgst'],tgst],1):
                cell=ws.cell(row=ri,column=ci,value=v); sc(cell,RF)
                if ci in [3,4,5,6,7]: cell.number_format='#,##0.00'
        tr=len(data)+3; ws.cell(row=tr,column=1,value="TOTAL").font=BLD
        for ci,key in [(4,'igst'),(5,'cgst'),(6,'sgst')]:
            c2=ws.cell(row=tr,column=ci,value=round(sum(r[key] for r in data),2))
            c2.font=BLD;c2.number_format='#,##0.00';c2.fill=LB;c2.border=BD
        ws.freeze_panes="A3"
        for i,w in enumerate([38,14,14,12,12,12,12],1): ws.column_dimensions[get_column_letter(i)].width=w

    mk_matched(wb,"✅ Matched Exact",res_exact,TH)
    mk_matched(wb,"✅ Matched (±10 Diff)",res_diff,OH,is_diff=True)
    mk_un2a(wb,res_un2a); mk_unb(wb,res_unb)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── RUN ────────────────────────────────────────────────────────────────────────
if run_btn:
    if not file_2a or not file_books:
        st.error("❌ Please upload BOTH files before running.")
    else:
        with st.spinner("🔄 Processing your files — please wait..."):
            try:
                b2a   = convert_xls_to_xlsx(file_2a.read(),    file_2a.name)
                bb    = convert_xls_to_xlsx(file_books.read(), file_books.name)
                wb2a  = openpyxl.load_workbook(b2a, read_only=True)
                wb_bk = openpyxl.load_workbook(bb,  read_only=True)
                gst2a = parse_2a(wb2a)
                books = parse_books(wb_bk)
                res_exact,res_diff,res_un2a,res_unb = run_matching(gst2a,books)
                out = build_excel(res_exact,res_diff,res_un2a,res_unb,gst2a,books)
                st.session_state.update({
                    "reco_done":True, "reco_out":out,
                    "res_exact":res_exact, "res_diff":res_diff,
                    "res_un2a":res_un2a, "res_unb":res_unb,
                    "gst2a_count":len(gst2a), "books_count":len(books),
                    "paid":False
                })
            except Exception as e:
                st.error(f"❌ Something went wrong: {str(e)}")
                st.exception(e)

# ── RESULTS + PAYWALL AT DOWNLOAD ONLY ────────────────────────────────────────
if st.session_state.get("reco_done"):
    re_ = st.session_state["res_exact"]
    rd_ = st.session_state["res_diff"]
    ru_ = st.session_state["res_un2a"]
    rb_ = st.session_state["res_unb"]

    # Summary shown FREE
    st.markdown(f"""
    <div class="stat-grid">
      <div class="stat-card"><div class="stat-number c-green">{len(re_)}</div><div class="stat-label">✅ Matched Exact</div></div>
      <div class="stat-card"><div class="stat-number c-yellow">{len(rd_)}</div><div class="stat-label">✅ Matched (±₹10)</div></div>
      <div class="stat-card"><div class="stat-number c-red">{len(ru_)}</div><div class="stat-label">⚠️ Unmatched in 2A</div></div>
    </div>
    <div class="stat-grid">
      <div class="stat-card"><div class="stat-number c-blue">{st.session_state["gst2a_count"]}</div><div class="stat-label">Total 2A Records</div></div>
      <div class="stat-card"><div class="stat-number c-blue">{st.session_state["books_count"]}</div><div class="stat-label">Total Books Records</div></div>
      <div class="stat-card"><div class="stat-number c-red">{len(rb_)}</div><div class="stat-label">⚠️ Unmatched in Books</div></div>
    </div>
    """, unsafe_allow_html=True)

    # Paywall only at download
    if not st.session_state.get("paid"):
        st.markdown(f"""
        <div class="pay-box">
          <h3>📥 Your report is ready! Pay ₹5 to download</h3>
          <p class="sub">Scan the QR · Pay ₹5 · Enter your UTR below · Download instantly</p>
          <div class="pay-inner">
            <div class="pay-left">
              <div class="pay-amount">₹5</div>
              <div class="pay-per">one-time · this reconciliation only</div>
              <div class="refund-tag">
                💚 <b>100% Refund Guarantee</b><br>
                Excel doesn't work? Send a screenshot — full refund, no questions asked.
              </div>
            </div>
            <div class="pay-right">
              <div class="qr-wrap">
                <img src="{QR_URL}" width="200" height="200" alt="Pay ₹5" />
              </div>
              <div class="qr-hint">Scan with any UPI app · ₹5 only</div>
              <div class="qr-apps">
                <span class="qr-app">GPay</span>
                <span class="qr-app">PhonePe</span>
                <span class="qr-app">Paytm</span>
                <span class="qr-app">BHIM</span>
              </div>
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        col_utr, col_unlock, _ = st.columns([2,1,1])
        with col_utr:
            utr = st.text_input("", placeholder="Enter UTR / Transaction ID after payment", label_visibility="collapsed")
        with col_unlock:
            if st.button("🔓 Unlock Download"):
                if len(utr.strip()) >= 8:
                    st.session_state["paid"] = True
                    st.rerun()
                else:
                    st.error("❌ Please enter a valid UTR (min 8 characters)")

    else:
        st.success("✅ Payment confirmed! Your Excel report is ready.")
        col_dl, _, _ = st.columns([1,1,1])
        with col_dl:
            st.download_button(
                "📥 Download Reconciliation Excel",
                data=st.session_state["reco_out"],
                file_name="GSTR2A_Reco_HarshMevada.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.markdown("<br>", unsafe_allow_html=True)
        col_reset, _, _ = st.columns([1,1,1])
        with col_reset:
            if st.button("🔄 Run Another Reconciliation"):
                for k in ["reco_done","reco_out","res_exact","res_diff","res_un2a","res_unb","paid","gst2a_count","books_count"]:
                    st.session_state.pop(k, None)
                st.rerun()

# ── FOOTER ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
  Built with ❤️ by <span>Harsh Mevada</span> · CA Tools India<br>
  GSTR-2A Reconciliation Tool · FY 2025-26<br>
  Fuzzy name matching · ±₹10 tolerance · Zero data storage · ₹5 per download
</div>
""", unsafe_allow_html=True)
